"""Builds the DefensiveIQ-style Excel workbook from an analyzed dataframe."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
from collections import Counter

from analyzer import (
    ZONE_ORDER, ZONE_LABEL, MIN_SAMPLE_CONCEPT, MIN_SAMPLE_FORMATION,
    SMALL_SAMPLE_FLAG, top_n_counts, run_pass_split, call_idea,
)

# ------------------------------------------------------------------
# Color palette — pulled directly from cell fills in a real DefensiveIQ
# workbook, not guessed. Reused consistently across every sheet below.
# ------------------------------------------------------------------
NAVY = "16213E"       # neutral sheet headers/titles
RUN_RED = "C0392B"    # run-themed titles/headers, 3rd-down urgency
PASS_BLUE = "1A5276"  # pass-themed titles/headers, formation group headers
PURPLE = "4A235A"     # goal line, hash/situational titles
TEAL = "0E7060"       # 1st down / "safe" situations, backed-up label
OLIVE = "7D6608"      # fringe zone
MAROON = "7B241C"     # 4th down (most critical)
YELLOW = "F1C40F"     # drive-start / call-sheet highlight
SKY_BLUE = "2E86C1"   # formation name labels

RUN_TINT = "FDE8E8"
PASS_TINT = "E8F0FE"
NEUTRAL_TINT_A = "F5F5F5"
NEUTRAL_TINT_B = "FFFFFF"

ZONE_STYLE = {
    "BZ": {"label": RUN_RED, "tint": "FDE8E8"},
    "OF": {"label": PASS_BLUE, "tint": "E8F0FE"},
    "MF": {"label": TEAL, "tint": "E8F8E8"},
    "FZ": {"label": OLIVE, "tint": "FFFBE6"},
    "RZ": {"label": RUN_RED, "tint": "FCE4EC"},
    "GL": {"label": PURPLE, "tint": "EDE7F6"},
}

# Tab colors, matched exactly to the sample workbook's own tab colors.
TAB_COLORS = {
    "1. Film Log": "0E7060",
    "2. Field Zone Tendencies": "1A5276",
    "3. Run Tendencies": "C0392B",
    "4. Pass Tendencies": "1A5276",
    "5. Hash Tendencies": "6C3483",
    "6. Down & Distance": "F1C40F",
    "7. Run Concepts": "8B0000",
    "8. Pass Concepts": "00008B",
    "9. Formation Tendencies": "2E86C1",
    "10. Situational Summary": "4A235A",
    "11. Practice Scripts": "2E86C1",
    "12. Call Sheet Builder": "C0392B",
    "13. Game Day Call Sheet": "F1C40F",
}

# Title-bar fill per sheet, matched to the sample workbook.
TITLE_FILLS = {
    "1. Film Log": TEAL,
    "2. Field Zone Tendencies": NAVY,
    "3. Run Tendencies": RUN_RED,
    "4. Pass Tendencies": PASS_BLUE,
    "5. Hash Tendencies": PURPLE,
    "6. Down & Distance": NAVY,
    "7. Run Concepts": NAVY,
    "8. Pass Concepts": NAVY,
    "9. Formation Tendencies": NAVY,
    "10. Situational Summary": PURPLE,
    "11. Practice Scripts": NAVY,
    "12. Call Sheet Builder": RUN_RED,
    "13. Game Day Call Sheet": NAVY,
}


def situation_color(label):
    """Down-progression urgency color for a situation/zone label cell,
    matching the pattern in the sample workbook: teal (1st/safe) -> blue
    (2nd) -> red (3rd) -> maroon (4th, most critical), with red zone,
    goal line, and drive-start situations carrying their own accent."""
    s = (label or "").upper()
    if "GOAL LINE" in s:
        return PURPLE
    if "RED ZONE" in s:
        return RUN_RED
    if "4TH" in s:
        return MAROON
    if "3RD" in s:
        return RUN_RED
    if "2ND" in s:
        return PASS_BLUE
    if "1ST" in s or "BACKED UP" in s or "COMING OUT" in s:
        return TEAL
    if "P & 10" in s or "DRIVE START" in s:
        return YELLOW
    return NAVY


def zebra_fill(row_index):
    return NEUTRAL_TINT_A if row_index % 2 == 0 else NEUTRAL_TINT_B


TITLE_FONT = Font(bold=True, size=13, color="FFFFFF", name="Arial")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BASE_FONT = Font(name="Arial", size=10)
RUN_FILL = PatternFill("solid", fgColor=RUN_TINT)
PASS_FILL = PatternFill("solid", fgColor=PASS_TINT)


def _tab_color(ws, sheet_name):
    ws.sheet_properties.tabColor = TAB_COLORS.get(sheet_name, NAVY)


def _title(ws, text, span=10, sheet_name=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = PatternFill("solid", fgColor=TITLE_FILLS.get(sheet_name, NAVY))
    ws.row_dimensions[1].height = 22
    if sheet_name:
        _tab_color(ws, sheet_name)


def _header_row(ws, row, headers, colors=None):
    """colors: optional list of per-column fill hex (same length as headers);
    falls back to navy for any column not overridden."""
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        fill_color = colors[i - 1] if colors and i - 1 < len(colors) and colors[i - 1] else NAVY
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _label_cell(ws, r, c, text, color, text_color="FFFFFF"):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = Font(bold=True, color=text_color, name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=color)
    return cell


def _autosize(ws, widths=None):
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 32)


def _write_row(ws, r, values, fills=None):
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = BASE_FONT
        if fills and i - 1 < len(fills) and fills[i - 1]:
            cell.fill = fills[i - 1]


# ------------------------------------------------------------------
# Sheet builders
# ------------------------------------------------------------------

def build_film_log(wb, df):
    ws = wb.create_sheet("1. Film Log")
    _title(ws, "FILM LOG  —  CONCEPT = play concept for ALL plays", span=13, sheet_name="1. Film Log")
    _header_row(ws, 2, ["QTR", "DN", "DIST", "HASH", "YARD LN", "ZONE", "OFF FORM",
                        "OFF STR", "BACKFIELD", "PLAY DIR", "PLAY TYPE", "CONCEPT", "GN/LS"],
                colors=[None, None, None, None, None, "000088"])
    r = 3
    for _, row in df.iterrows():
        base_fill = PatternFill("solid", fgColor=zebra_fill(r))
        zone = row.get("ZONE")
        zone_tint = ZONE_STYLE.get(zone, {}).get("tint", "E8F4FD")
        fills = [base_fill] * 5 + [PatternFill("solid", fgColor=zone_tint)] + [base_fill] * 7
        _write_row(ws, r, [
            None, row.get("DN"), row.get("DIST"), row.get("HASH"), None,
            row.get("ZONE"), row.get("FORMATION"), None, row.get("BACKFIELD"),
            None, row.get("PLAY TYPE"), row.get("CONCEPT"), row.get("GAIN"),
        ], fills=fills)
        r += 1
    _autosize(ws)


def build_field_zone_tendencies(wb, df):
    ws = wb.create_sheet("2. Field Zone Tendencies")
    _title(ws, "FIELD ZONE TENDENCIES", span=10, sheet_name="2. Field Zone Tendencies")
    ws.cell(row=2, column=1, value="  Gray=plays  Red%=Run  Blue%=Pass  Yellow=Call Idea").font = Font(italic=True, size=9)
    ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor=NEUTRAL_TINT_A)
    dd_cols = ["1st Down", "2nd & 7+", "2nd & 4-6", "2nd & 1-3", "3rd & 7+", "3rd & 4-6", "3rd & 1-3", "4th Down"]
    dd_map = {
        "1st Down": lambda s: s["DN"] == 1,
        "2nd & 7+": lambda s: (s["DN"] == 2) & (s["DIST"] >= 7),
        "2nd & 4-6": lambda s: (s["DN"] == 2) & (s["DIST"].between(4, 6)),
        "2nd & 1-3": lambda s: (s["DN"] == 2) & (s["DIST"].between(1, 3)),
        "3rd & 7+": lambda s: (s["DN"] == 3) & (s["DIST"] >= 7),
        "3rd & 4-6": lambda s: (s["DN"] == 3) & (s["DIST"].between(4, 6)),
        "3rd & 1-3": lambda s: (s["DN"] == 3) & (s["DIST"].between(1, 3)),
        "4th Down": lambda s: s["DN"] == 4,
    }
    _header_row(ws, 3, ["FIELD ZONE", "METRIC"] + dd_cols)
    r = 4
    for zone in ZONE_ORDER:
        zdf = df[df["ZONE"] == zone]
        zone_color = ZONE_STYLE[zone]["label"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        _label_cell(ws, r, 1, f"  {zone}  ·  {ZONE_LABEL[zone]}  ({len(zdf)} plays)", zone_color)
        r += 1
        plays_row, run_row, pass_row, runs_row, passes_row = [], [], [], [], []
        for col in dd_cols:
            mask = dd_map[col](zdf)
            sub = zdf[mask]
            runs, passes, total, run_pct, pass_pct = run_pass_split(sub)
            plays_row.append(total if total else 0)
            run_row.append(round(run_pct, 2) if run_pct is not None else None)
            pass_row.append(round(pass_pct, 2) if pass_pct is not None else None)
            runs_row.append(runs)
            passes_row.append(passes)
        gray = PatternFill("solid", fgColor=NEUTRAL_TINT_A)
        _write_row(ws, r, [None, "Plays"] + plays_row, fills=[gray, gray] + [gray] * len(dd_cols)); r += 1
        _write_row(ws, r, [None, "Run %"] + run_row, fills=[None, None] + [RUN_FILL] * len(dd_cols)); r += 1
        _write_row(ws, r, [None, "Pass %"] + pass_row, fills=[None, None] + [PASS_FILL] * len(dd_cols)); r += 1
        _write_row(ws, r, [None, "Runs"] + runs_row, fills=[gray, gray] + [gray] * len(dd_cols)); r += 1
        _write_row(ws, r, [None, "Passes"] + passes_row, fills=[gray, gray] + [gray] * len(dd_cols)); r += 1
        runs_z, passes_z, total_z, run_pct_z, pass_pct_z = run_pass_split(zdf)
        yellow_fill = PatternFill("solid", fgColor=YELLOW)
        _write_row(ws, r, [None, "▶ Call Idea", call_idea(run_pct_z, pass_pct_z)], fills=[None, yellow_fill, yellow_fill])
        r += 2
    _autosize(ws, widths=[32, 10] + [11] * 8)
    # percent formatting
    for row in ws.iter_rows(min_row=4, min_col=3, max_col=10):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0%"


def build_run_pass_tendencies(wb, df, play_type, sheet_name, expl_yards, top_col3_name, top_col3_short, top_col3_source):
    ws = wb.create_sheet(sheet_name)
    _title(ws, f"{play_type.upper()} TENDENCIES  —  Auto-calculated from Hudl data", span=13, sheet_name=sheet_name)
    formation_header_color = PASS_BLUE
    _header_row(ws, 2, ["ZONE", "COUNTS", "", "", "TOP FORMATIONS", "", "", f"TOP {play_type.upper()} CONCEPTS", "", "", top_col3_name, "", ""],
                colors=[None, None, None, None, formation_header_color, formation_header_color, formation_header_color])
    _header_row(ws, 3, ["Zone", f"{play_type}\nCount", f"{play_type} %", "Explosive\n" + play_type + "s",
                        "#1 Formation", "#2 Formation", "#3 Formation",
                        "#1 Concept", "#2 Concept", "#3 Concept",
                        f"#1 {top_col3_short}", f"#2 {top_col3_short}", f"#3 {top_col3_short}"],
                colors=[None, None, None, None, formation_header_color, formation_header_color, formation_header_color])
    r = 4
    for zone in ZONE_ORDER:
        zdf = df[(df["ZONE"] == zone) & (df["PLAY TYPE"] == play_type)]
        all_zdf = df[df["ZONE"] == zone]
        count = len(zdf)
        pct_ = count / len(all_zdf) if len(all_zdf) else 0
        expl = zdf["EXPLOSIVE"].sum()
        forms = top_n_counts(zdf["FORMATION"], 3)
        concepts = top_n_counts(zdf["CONCEPT"], 3)
        col3 = top_n_counts(zdf[top_col3_source], 3) if count else ["—", "—", "—"]
        zone_tint = PatternFill("solid", fgColor=ZONE_STYLE[zone]["tint"])
        _label_cell(ws, r, 1, zone, ZONE_STYLE[zone]["label"])
        _write_row(ws, r, [None, count, round(pct_, 2), int(expl)] + forms + concepts + col3,
                   fills=[None] + [zone_tint] * 12)
        r += 1
    for row in ws.iter_rows(min_row=4, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "0%"
    _autosize(ws)


def build_hash_tendencies(wb, df):
    ws = wb.create_sheet("5. Hash Tendencies")
    _title(ws, "HASH TENDENCIES  —  Left Hash · Middle · Right Hash", span=13, sheet_name="5. Hash Tendencies")
    _header_row(ws, 2, ["FIELD ZONE", "L Plays", "L Run%", "L Pass%", "M Plays", "M Run%", "M Pass%",
                        "R Plays", "R Run%", "R Pass%", "Top L Concept", "Top M Concept", "Top R Concept"])

    def hash_block(sub):
        out = []
        for h in ["L", "M", "R"]:
            hsub = sub[sub["HASH"] == h]
            runs, passes, total, run_pct, pass_pct = run_pass_split(hsub)
            out += [total, round(run_pct, 2) if run_pct is not None else None,
                    round(pass_pct, 2) if pass_pct is not None else None]
        concepts = []
        for h in ["L", "M", "R"]:
            hsub = sub[sub["HASH"] == h]
            top = top_n_counts(hsub["CONCEPT"], 1)[0] if len(hsub) else "—"
            concepts.append(top)
        return out, concepts

    gray = PatternFill("solid", fgColor=NEUTRAL_TINT_A)
    row_fills = [gray, RUN_FILL, PASS_FILL, gray, RUN_FILL, PASS_FILL, gray, RUN_FILL, PASS_FILL, None, None, None]

    r = 3
    out, concepts = hash_block(df)
    _label_cell(ws, r, 1, "OVERALL", NAVY)
    _write_row(ws, r, [None] + out + concepts, fills=[None] + row_fills)
    r += 1
    for zone in ZONE_ORDER:
        zdf = df[df["ZONE"] == zone]
        out, concepts = hash_block(zdf)
        _label_cell(ws, r, 1, f"{zone} — {ZONE_LABEL[zone]}", ZONE_STYLE[zone]["label"])
        _write_row(ws, r, [None] + out + concepts, fills=[None] + row_fills)
        r += 1
    for row in ws.iter_rows(min_row=3, min_col=3, max_col=10):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column in (3, 4, 6, 7, 9, 10):
                cell.number_format = "0%"
    _autosize(ws, widths=[26] + [9] * 9 + [18, 18, 18])


DD_SITUATIONS = ["1ST & 10", "1ST & SHORT", "2ND & LONG", "2ND & MEDIUM", "2ND & SHORT",
                 "3RD & LONG", "3RD & MEDIUM", "3RD & SHORT", "4TH DOWN"]
ZONE_SITUATIONS = [("RED ZONE", "RZ"), ("GOAL LINE", "GL"), ("BACKED UP", "BZ")]


def build_down_distance(wb, df):
    ws = wb.create_sheet("6. Down & Distance")
    _title(ws, "DOWN & DISTANCE TENDENCIES  —  Favorite Runs, Passes & Formations by Situation", span=13, sheet_name="6. Down & Distance")
    _header_row(ws, 2, ["SITUATION", "Plays", "Run%", "Pass%", "#1 Run Concept", "#2 Run Concept", "#3 Run Concept",
                        "#1 Pass Concept", "#2 Pass Concept", "#3 Pass Concept", "#1 Formation", "#2 Formation", "#3 Formation"],
                colors=[None, None, None, None, RUN_RED, RUN_RED, RUN_RED, PASS_BLUE, PASS_BLUE, PASS_BLUE, SKY_BLUE, SKY_BLUE, SKY_BLUE])
    r = 3

    def write_situation(label, sub):
        nonlocal r
        runs, passes, total, run_pct, pass_pct = run_pass_split(sub)
        run_concepts = top_n_counts(sub[sub["PLAY TYPE"] == "Run"]["CONCEPT"], 3)
        pass_concepts = top_n_counts(sub[sub["PLAY TYPE"] == "Pass"]["CONCEPT"], 3)
        forms = top_n_counts(sub["FORMATION"], 3)
        gray = PatternFill("solid", fgColor=zebra_fill(r))
        _label_cell(ws, r, 1, label, situation_color(label))
        _write_row(ws, r, [None, total, round(run_pct, 2) if run_pct is not None else None,
                           round(pass_pct, 2) if pass_pct is not None else None] + run_concepts + pass_concepts + forms,
                   fills=[None, gray, RUN_FILL, PASS_FILL] + [RUN_FILL] * 3 + [PASS_FILL] * 3 + [gray] * 3)
        r += 1

    for sit in DD_SITUATIONS:
        write_situation(sit, df[df["DD_BUCKET"] == sit])
    for label, zone in ZONE_SITUATIONS:
        write_situation(label, df[df["ZONE"] == zone])
    for row in ws.iter_rows(min_row=3, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = "0%"
    _autosize(ws)


def build_concepts(wb, df, play_type, sheet_name, expl_yards):
    ws = wb.create_sheet(sheet_name)
    accent = RUN_RED if play_type == "Run" else PASS_BLUE
    tint_fill = RUN_FILL if play_type == "Run" else PASS_FILL
    _title(ws, f"{play_type.upper()} CONCEPTS  —  Explosive ({expl_yards}+) & Success Rate by Concept ({MIN_SAMPLE_CONCEPT}+ calls)",
           span=13, sheet_name=sheet_name)
    _header_row(ws, 2, [f"{play_type.upper()} CONCEPT", "Called", "Avg Yd", "Expl%", "Succ%", "Dir R%", "Dir L%",
                        "1st Dn Succ", "2nd Dn Succ", "3rd Dn Succ", "RedZone Succ", "Top Form", "Top Hash"],
                colors=[accent, None, None, accent, accent, accent, accent, accent, accent, accent, accent, None, None])
    sub = df[df["PLAY TYPE"] == play_type]
    counts = sub["CONCEPT"].value_counts()
    r = 3
    for concept, called in counts.items():
        if called < MIN_SAMPLE_CONCEPT or pd.isna(concept):
            continue
        cdf = sub[sub["CONCEPT"] == concept]
        avg_yd = cdf["GAIN"].mean()
        expl_pct = cdf["EXPLOSIVE"].mean()
        succ_series = cdf["SUCCESS"].dropna()
        succ_pct = succ_series.mean() if len(succ_series) else None

        def succ_by_down(dn):
            d = cdf[cdf["DN"] == dn]["SUCCESS"].dropna()
            return round(d.mean(), 2) if len(d) else "—"

        rz = cdf[cdf["ZONE"] == "RZ"]["SUCCESS"].dropna()
        rz_succ = round(rz.mean(), 2) if len(rz) else "—"
        top_form = top_n_counts(cdf["FORMATION"], 1)[0].split(" (")[0]
        top_hash = cdf["HASH"].mode().iloc[0] if not cdf["HASH"].mode().empty else "—"
        label = f"{concept} *" if called < SMALL_SAMPLE_FLAG else concept
        gray = PatternFill("solid", fgColor=zebra_fill(r))
        _label_cell(ws, r, 1, label, accent)
        _write_row(ws, r, [
            None, int(called), round(avg_yd, 1) if pd.notna(avg_yd) else None,
            round(expl_pct, 2), round(succ_pct, 2) if succ_pct is not None else None,
            None, None, succ_by_down(1), succ_by_down(2), succ_by_down(3), rz_succ, top_form, top_hash,
        ], fills=[None, gray, gray, tint_fill, tint_fill, None, None, tint_fill, tint_fill, tint_fill, tint_fill, gray, gray])
        r += 1
    ws.cell(row=r + 1, column=1, value=f"* = small sample (under {SMALL_SAMPLE_FLAG} calls) — read with caution").font = Font(italic=True, size=9)
    for row in ws.iter_rows(min_row=3, min_col=4, max_col=5):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0%"
    for row in ws.iter_rows(min_row=3, min_col=8, max_col=11):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0%"
    _autosize(ws)


def build_formation_tendencies(wb, df):
    ws = wb.create_sheet("9. Formation Tendencies")
    _title(ws, f"FORMATION TENDENCIES  —  Favorite Runs & Passes by Formation ({MIN_SAMPLE_FORMATION}+ snaps)", span=12, sheet_name="9. Formation Tendencies")
    _header_row(ws, 2, ["FORMATION", "Snaps", "Run%", "Pass%", "Run R%", "Run L%",
                        "#1 Run Concept", "#2 Run Concept", "#1 Pass Concept", "#2 Pass Concept",
                        "Top Down/Dist", "Top Zone"],
                colors=[None, None, None, None, None, None, RUN_RED, RUN_RED, PASS_BLUE, PASS_BLUE])
    counts = df["FORMATION"].value_counts()
    r = 3
    for form, snaps in counts.items():
        if snaps < MIN_SAMPLE_FORMATION or pd.isna(form):
            continue
        fdf = df[df["FORMATION"] == form]
        runs, passes, total, run_pct, pass_pct = run_pass_split(fdf)
        run_concepts = top_n_counts(fdf[fdf["PLAY TYPE"] == "Run"]["CONCEPT"], 2)
        pass_concepts = top_n_counts(fdf[fdf["PLAY TYPE"] == "Pass"]["CONCEPT"], 2)
        top_dd = top_n_counts(fdf["DD_BUCKET"], 1)[0]
        top_zone_code = fdf["ZONE"].mode().iloc[0] if not fdf["ZONE"].mode().empty else None
        top_zone_count = (fdf["ZONE"] == top_zone_code).sum() if top_zone_code else 0
        top_zone = f"{ZONE_LABEL.get(top_zone_code,'—').split('·')[0].strip()} ({top_zone_count})" if top_zone_code else "—"
        gray = PatternFill("solid", fgColor=zebra_fill(r))
        _label_cell(ws, r, 1, form, SKY_BLUE)
        _write_row(ws, r, [None, int(snaps), round(run_pct, 2), round(pass_pct, 2), None, None] +
                   run_concepts + pass_concepts + [top_dd, top_zone],
                   fills=[None, gray, RUN_FILL, PASS_FILL, gray, gray, RUN_FILL, RUN_FILL, PASS_FILL, PASS_FILL, gray, gray])
        r += 1
    for row in ws.iter_rows(min_row=3, min_col=3, max_col=4):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0%"
    _autosize(ws)


SITUATIONAL_ROWS = [
    ("P & 10 (Drive Start)", lambda df: df[df["DN"] == 0]),
    ("1ST DOWN", lambda df: df[df["DN"] == 1]),
    ("2ND & LONG", lambda df: df[df["DD_BUCKET"] == "2ND & LONG"]),
    ("2ND & MEDIUM", lambda df: df[df["DD_BUCKET"] == "2ND & MEDIUM"]),
    ("2ND & SHORT", lambda df: df[df["DD_BUCKET"] == "2ND & SHORT"]),
    ("3RD & LONG", lambda df: df[df["DD_BUCKET"] == "3RD & LONG"]),
    ("3RD & MEDIUM", lambda df: df[df["DD_BUCKET"] == "3RD & MEDIUM"]),
    ("3RD & SHORT", lambda df: df[df["DD_BUCKET"] == "3RD & SHORT"]),
    ("4TH DOWN", lambda df: df[df["DN"] == 4]),
    ("RED ZONE", lambda df: df[df["ZONE"] == "RZ"]),
    ("GOAL LINE", lambda df: df[df["ZONE"] == "GL"]),
    ("BACKED UP", lambda df: df[df["ZONE"] == "BZ"]),
    ("COMING OUT", lambda df: df[df["ZONE"].isin(["BZ", "OF"])]),
]


def build_situational_summary(wb, df):
    ws = wb.create_sheet("10. Situational Summary")
    _title(ws, "SITUATIONAL SUMMARY", span=10, sheet_name="10. Situational Summary")
    _header_row(ws, 2, ["Situation", "Run\nCount", "Pass\nCount", "L Hash\nRun%", "M Hash\nRun%",
                        "R Hash\nRun%", "Top Run Concept", "Top Pass Concept", "Best Call", "Notes"])
    r = 3
    for label, fn in SITUATIONAL_ROWS:
        sub = fn(df)
        runs = (sub["PLAY TYPE"] == "Run").sum()
        passes = (sub["PLAY TYPE"] == "Pass").sum()

        def hash_run_pct(h):
            hsub = sub[sub["HASH"] == h]
            r_, p_, t_, rp, pp = run_pass_split(hsub)
            return round(rp, 2) if rp is not None else None

        top_run = top_n_counts(sub[sub["PLAY TYPE"] == "Run"]["CONCEPT"], 1)[0]
        top_pass = top_n_counts(sub[sub["PLAY TYPE"] == "Pass"]["CONCEPT"], 1)[0]
        gray = PatternFill("solid", fgColor=zebra_fill(r))
        _label_cell(ws, r, 1, label, situation_color(label))
        _write_row(ws, r, [None, int(runs), int(passes), hash_run_pct("L"), hash_run_pct("M"), hash_run_pct("R"),
                           top_run, top_pass, None, None],
                   fills=[None, RUN_FILL, PASS_FILL, gray, gray, gray, gray, gray, None, None])
        r += 1
    # No live-clock data in a Hudl playlist export, so two-minute / must-have
    # situations can't be isolated - mirror overall totals as the original tool does.
    all_runs = (df["PLAY TYPE"] == "Run").sum()
    all_passes = (df["PLAY TYPE"] == "Pass").sum()
    top_run = top_n_counts(df[df["PLAY TYPE"] == "Run"]["CONCEPT"], 1)[0]
    top_pass = top_n_counts(df[df["PLAY TYPE"] == "Pass"]["CONCEPT"], 1)[0]
    for label in ["TWO-MINUTE", "MUST-HAVE"]:
        gray = PatternFill("solid", fgColor=zebra_fill(r))
        _label_cell(ws, r, 1, label, NAVY)
        _write_row(ws, r, [None, int(all_runs), int(all_passes), None, None, None, top_run, top_pass, None,
                           "No clock data in export - shown as overall totals"],
                   fills=[None, RUN_FILL, PASS_FILL, gray, gray, gray, gray, gray, None, None])
        r += 1
    for row in ws.iter_rows(min_row=3, min_col=4, max_col=6):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0%"
    _autosize(ws)


def build_practice_scripts(wb, df):
    ws = wb.create_sheet("11. Practice Scripts")
    _title(ws, "PRACTICE SCRIPTS   ·   auto-pulled from film", span=8, sheet_name="11. Practice Scripts")
    ws.cell(row=3, column=1, value="MONDAY  ·  FAVORITES (ALL DOWNS)").font = Font(bold=True)
    section_fill = PatternFill("solid", fgColor=NAVY)
    c = ws.cell(row=4, column=1, value="FORMATION ALIGNMENT  (top 12)")
    c.font = Font(bold=True, italic=True, color="FFFFFF")
    c.fill = section_fill
    _header_row(ws, 5, ["#", "HASH", "FORMATION", "PERS", "FRONT", "COVERAGE"])
    counts = df["FORMATION"].value_counts().head(12)
    r = 6
    for i, (form, cnt) in enumerate(counts.items(), start=1):
        fdf = df[df["FORMATION"] == form]
        top_hash = fdf["HASH"].mode().iloc[0] if not fdf["HASH"].mode().empty else "—"
        gray = PatternFill("solid", fgColor=zebra_fill(r))
        _write_row(ws, r, [i, top_hash, form, None, None, None], fills=[gray] * 6)
        r += 1
    r += 1
    run_fill_hdr = PatternFill("solid", fgColor=RUN_RED)
    c = ws.cell(row=r, column=1, value="INSIDE SCRIPT  (top run concepts, coach fills front/coverage)")
    c.font = Font(bold=True, italic=True, color="FFFFFF")
    c.fill = run_fill_hdr
    r += 1
    _header_row(ws, r, ["#", "HASH", "RUN CONCEPT", "FORMATION", "FRONT", "COVERAGE"], colors=[RUN_RED] * 6)
    r += 1
    run_counts = df[df["PLAY TYPE"] == "Run"]["CONCEPT"].value_counts().head(12)
    for i, (concept, cnt) in enumerate(run_counts.items(), start=1):
        cdf = df[(df["PLAY TYPE"] == "Run") & (df["CONCEPT"] == concept)]
        top_hash = cdf["HASH"].mode().iloc[0] if not cdf["HASH"].mode().empty else "—"
        top_form = top_n_counts(cdf["FORMATION"], 1)[0].split(" (")[0]
        _write_row(ws, r, [i, top_hash, concept, top_form, None, None], fills=[RUN_FILL] * 6)
        r += 1
    r += 1
    pass_fill_hdr = PatternFill("solid", fgColor=PASS_BLUE)
    c = ws.cell(row=r, column=1, value="SKELETON / PASS SCRIPT  (top pass concepts)")
    c.font = Font(bold=True, italic=True, color="FFFFFF")
    c.fill = pass_fill_hdr
    r += 1
    _header_row(ws, r, ["#", "HASH", "PASS CONCEPT", "FORMATION", "FRONT", "COVERAGE"], colors=[PASS_BLUE] * 6)
    r += 1
    pass_counts = df[df["PLAY TYPE"] == "Pass"]["CONCEPT"].value_counts().head(12)
    for i, (concept, cnt) in enumerate(pass_counts.items(), start=1):
        cdf = df[(df["PLAY TYPE"] == "Pass") & (df["CONCEPT"] == concept)]
        top_hash = cdf["HASH"].mode().iloc[0] if not cdf["HASH"].mode().empty else "—"
        top_form = top_n_counts(cdf["FORMATION"], 1)[0].split(" (")[0]
        _write_row(ws, r, [i, top_hash, concept, top_form, None, None], fills=[PASS_FILL] * 6)
        r += 1
    _autosize(ws)


CALL_SHEET_SITUATIONS = [
    "1st & 10", "1st & 10 (Own Half)", "1st & 10 (Opp Half)", "2nd & Long (8+)",
    "2nd & Medium (4-7)", "2nd & Short (1-3)", "3rd & Long (7+)", "3rd & Medium (4-6)",
    "3rd & Short (1-3)", "4th Down", "Red Zone — 1st", "Red Zone — 2nd", "Red Zone — 3rd",
    "Goal Line", "Backed Up", "Coming Out", "Two-Minute (Lead)", "Two-Minute (Trail)",
    "Must-Have Plays", "Two-Point Play", "Overtime",
]


def build_call_sheet_builder(wb):
    ws = wb.create_sheet("12. Call Sheet Builder")
    _title(ws, "CALL SHEET BUILDER  —  Fill in your calls", span=11, sheet_name="12. Call Sheet Builder")
    _header_row(ws, 2, ["Situation", "Field\nZone", "Down /\nDistance", "Opponent Tendency", "Formation / Alert",
                        "Best Pressure", "Best Coverage", "Best Front", "Adjustment", "Priority", "Notes"],
                colors=[RUN_RED] * 11)
    r = 3
    for i, sit in enumerate(CALL_SHEET_SITUATIONS):
        pink_zebra = "FFF0F0" if i % 2 == 0 else "FFFFFF"
        gray = PatternFill("solid", fgColor=pink_zebra)
        _write_row(ws, r, [sit] + [None] * 10, fills=[gray] * 11)
        r += 1
    _autosize(ws)


def build_game_day_call_sheet(wb, df, opponent, week):
    ws = wb.create_sheet("13. Game Day Call Sheet")
    _title(ws, f"GAME DAY CALL SHEET   ·   {opponent}   ·   WK {week}", span=11, sheet_name="13. Game Day Call Sheet")
    _header_row(ws, 3, ["Situation", "Run%", "Pass%", "Top Run", "Top Pass", "Form", "", "BIGGEST TENDENCIES", "", "", ""],
                colors=[None, RUN_RED, PASS_BLUE, None, None, None, None, PASS_BLUE])

    dd_labels = [("1st & 10", "1ST & 10"), ("2nd & Short", "2ND & SHORT"), ("2nd & Med", "2ND & MEDIUM"),
                 ("2nd & Long", "2ND & LONG"), ("3rd & Short", "3RD & SHORT"), ("3rd & Med", "3RD & MEDIUM"),
                 ("3rd & Long", "3RD & LONG"), ("4th Down", "4TH DOWN")]
    r = 4
    for label, bucket in dd_labels:
        sub = df[df["DD_BUCKET"] == bucket]
        runs, passes, total, run_pct, pass_pct = run_pass_split(sub)
        top_run = top_n_counts(sub[sub["PLAY TYPE"] == "Run"]["CONCEPT"], 1)[0].split(" (")[0]
        top_pass = top_n_counts(sub[sub["PLAY TYPE"] == "Pass"]["CONCEPT"], 1)[0].split(" (")[0]
        top_form = top_n_counts(sub["FORMATION"], 1)[0].split(" (")[0]
        gray = PatternFill("solid", fgColor=zebra_fill(r))
        _label_cell(ws, r, 1, label, situation_color(bucket))
        _write_row(ws, r, [None, round(run_pct, 2) if run_pct is not None else None,
                           round(pass_pct, 2) if pass_pct is not None else None, top_run, top_pass, top_form],
                   fills=[None, RUN_FILL, PASS_FILL, gray, gray, gray])
        r += 1

    r += 1
    zone_hdr_row = r
    c = ws.cell(row=r, column=1, value="FIELD ZONE")
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    r += 1
    _header_row(ws, r, ["Zone", "Run%", "Pass%", "Top Run", "Top Pass", "Form"], colors=[None, RUN_RED, PASS_BLUE])
    r += 1
    zone_labels = [("Coming Out", ["BZ"]), ("Open Field", ["OF"]), ("Midfield", ["MF"]),
                   ("Fringe", ["FZ"]), ("Red Zone", ["RZ"]), ("Goal Line", ["GL"])]
    for label, zones in zone_labels:
        sub = df[df["ZONE"].isin(zones)]
        runs, passes, total, run_pct, pass_pct = run_pass_split(sub)
        top_run = top_n_counts(sub[sub["PLAY TYPE"] == "Run"]["CONCEPT"], 1)[0].split(" (")[0]
        top_pass = top_n_counts(sub[sub["PLAY TYPE"] == "Pass"]["CONCEPT"], 1)[0].split(" (")[0]
        top_form = top_n_counts(sub["FORMATION"], 1)[0].split(" (")[0]
        gray = PatternFill("solid", fgColor=zebra_fill(r))
        _label_cell(ws, r, 1, label, ZONE_STYLE[zones[0]]["label"])
        _write_row(ws, r, [None, round(run_pct, 2) if run_pct is not None else None,
                           round(pass_pct, 2) if pass_pct is not None else None, top_run, top_pass, top_form],
                   fills=[None, RUN_FILL, PASS_FILL, gray, gray, gray])
        r += 1

    # Right column: biggest tendencies (formations with most lopsided run/pass split, min sample)
    tendencies = []
    for form, cnt in df["FORMATION"].value_counts().items():
        if cnt < MIN_SAMPLE_FORMATION or pd.isna(form):
            continue
        fdf = df[df["FORMATION"] == form]
        runs, passes, total, run_pct, pass_pct = run_pass_split(fdf)
        skew = max(run_pct, pass_pct)
        tag = "Run" if run_pct >= pass_pct else "Pass"
        flag = " *" if cnt < SMALL_SAMPLE_FLAG else ""
        tendencies.append((skew, f"{form} = {round(skew*100)}% {tag}{flag}   [{cnt}]"))
    tendencies.sort(reverse=True)
    tend_fill = PatternFill("solid", fgColor="E8F0FE")
    _write_row(ws, 4, [None] * 7 + ["#1", tendencies[0][1] if tendencies else "—"], fills=[None] * 7 + [tend_fill, tend_fill])
    for i in range(1, min(5, len(tendencies))):
        _write_row(ws, 4 + i, [None] * 7 + [f"#{i+1}", tendencies[i][1]], fills=[None] * 7 + [tend_fill, tend_fill])

    # Heavy pass situations
    heavy_row = 4 + 8
    c = ws.cell(row=heavy_row, column=8, value="HEAVY PASS SITUATIONS")
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=PASS_BLUE)
    heavy = []
    for label, fn in [("1st & 10", lambda d: d[d["DD_BUCKET"] == "1ST & 10"]),
                       ("2nd & Long", lambda d: d[d["DD_BUCKET"] == "2ND & LONG"]),
                       ("3rd & Long", lambda d: d[d["DD_BUCKET"] == "3RD & LONG"]),
                       ("3rd & Medium", lambda d: d[d["DD_BUCKET"] == "3RD & MEDIUM"]),
                       ("3rd & Short", lambda d: d[d["DD_BUCKET"] == "3RD & SHORT"]),
                       ("Backed Up", lambda d: d[d["ZONE"] == "BZ"])]:
        sub = fn(df)
        runs, passes, total, run_pct, pass_pct = run_pass_split(sub)
        if pass_pct is not None and total:
            heavy.append((pass_pct, f"{label}: {round(pass_pct*100)}% Pass" + (" *" if total < SMALL_SAMPLE_FLAG else "")))
    heavy.sort(reverse=True)
    for i, (pp, text) in enumerate(heavy[:5], start=1):
        _write_row(ws, heavy_row + i, [None] * 7 + [f"#{i}", text], fills=[None] * 7 + [PASS_FILL, PASS_FILL])

    # Red zone cheat sheet
    rz_row = heavy_row + 7
    c = ws.cell(row=rz_row, column=8, value="RED ZONE CHEAT SHEET")
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=RUN_RED)
    rzdf = df[df["ZONE"] == "RZ"]
    runs, passes, total, run_pct, pass_pct = run_pass_split(rzdf)
    top_run = top_n_counts(rzdf[rzdf["PLAY TYPE"] == "Run"]["CONCEPT"], 1)[0].split(" (")[0]
    top_pass = top_n_counts(rzdf[rzdf["PLAY TYPE"] == "Pass"]["CONCEPT"], 1)[0].split(" (")[0]
    top_form = top_n_counts(rzdf["FORMATION"], 1)[0].split(" (")[0]
    gldf = df[df["ZONE"] == "GL"]
    gl_top = top_n_counts(gldf["CONCEPT"], 1)[0] if len(gldf) else "—"
    rows = [
        ("Inside 20", f"Run {round((run_pct or 0)*100)}% / Pass {round((pass_pct or 0)*100)}%  ({total} plays)"),
        ("Top Run", top_run),
        ("Top Pass", top_pass),
        ("Top Formation", top_form),
        ("Goal Line Favorite", gl_top),
    ]
    rz_fill = PatternFill("solid", fgColor="FCE4EC")
    for i, (label, val) in enumerate(rows, start=1):
        _write_row(ws, rz_row + i, [None] * 7 + [label, val], fills=[None] * 7 + [rz_fill, rz_fill])

    for row in ws.iter_rows(min_row=4, min_col=2, max_col=3):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0%"

    footer_row = rz_row + 8
    ws.cell(row=footer_row, column=1,
            value="* = small sample (under 5 plays). Auto-generated from uploaded film.").font = Font(italic=True, size=9)
    _autosize(ws, widths=[18, 9, 9, 16, 16, 16, 3, 16, 30, 6, 6])


def build_workbook(df, opponent="Opponent", week="1"):
    wb = Workbook()
    wb.remove(wb.active)
    build_film_log(wb, df)
    build_field_zone_tendencies(wb, df)
    build_run_pass_tendencies(wb, df, "Run", "3. Run Tendencies", 10, "TOP HASH (no true direction data in export)", "Hash", "HASH")
    build_run_pass_tendencies(wb, df, "Pass", "4. Pass Tendencies", 15, "TOP BACKFIELD", "Backfield", "BACKFIELD")
    build_hash_tendencies(wb, df)
    build_down_distance(wb, df)
    build_concepts(wb, df, "Run", "7. Run Concepts", 10)
    build_concepts(wb, df, "Pass", "8. Pass Concepts", 15)
    build_formation_tendencies(wb, df)
    build_situational_summary(wb, df)
    build_practice_scripts(wb, df)
    build_call_sheet_builder(wb)
    build_game_day_call_sheet(wb, df, opponent, week)
    return wb
